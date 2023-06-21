import csv
import json
from datetime import datetime
import openpyxl
from openpyxl import Workbook  # for xlsx file
from bs4 import BeautifulSoup as bs  # for HTML table
import re  # for removing HTML tags
from fpdf import FPDF  # for creating PDF table
import PyPDF2  # for reading PDF table


# .txt to CSV

# Create txt file
with open('task4.txt', 'w') as f:
    f.write('20201201, Almaty, 130000')

# Convert to CSV and change date format
with open('task4.txt', 'r') as in_file, open('task4_fromtxt_out.csv', 'w') as out_file:
    reader = in_file.readlines()
    writer = csv.writer(out_file)
    data = [item.strip() for item in reader[0].split(',')]

    ts = datetime.strptime(data[0], "%Y%m%d")
    data[0] = ts.strftime('%d-%m-%Y')

    writer.writerow(('date', 'place', 'value'))
    writer.writerow(data)


# JSON to CSV

# Create JSON file
dictionary = {"date": "20201201", "place": "Almaty", "value": 130000}

json_object = json.dumps(dictionary)

with open("task4.json", "w") as out_file:
    out_file.write(json_object)

# Convert to CSV and change date format
with open("task4.json", "r") as in_file, open("task4_fromjson_out.csv", "w") as out_file:
    data = json.load(in_file)
    writer = csv.writer(out_file)

    ts = datetime.strptime(data['date'], "%Y%m%d")
    data['date'] = ts.strftime('%d-%m-%Y')
    writer.writerow(('date', 'place', 'value'))
    writer.writerow([data['date'], data['place'], data['value']])


# XLSX to CSV
wb = Workbook()
ws = wb.active

ws['A1'] = 'date'
ws['B1'] = 'place'
ws['C1'] = 'value'
ws['A2'] = '20201201'
ws['B2'] = 'Almaty'
ws['C2'] = 13000

wb.save(filename="task4.xlsx")

# Convert XLSX to CSV and change date format
excel = openpyxl.load_workbook("task4.xlsx", data_only=True)
sheet = excel.active

with open('task4_fromxlsx_out.csv', 'w') as out_file:
    writer = csv.writer(out_file)
    excel_date = sheet['A2'].value
    ts = datetime.strptime(excel_date, "%Y%m%d")
    excel_date = ts.strftime('%d-%m-%Y')
    sheet['A2'] = excel_date
    for row in sheet.rows:
        writer.writerow([cell.value for cell in row])


# HTML table to CSV

# Create HTML table
table = '''<table>
              <tr>
                <th>date</th>
                <th>place</th>
                <th>value</th>
              </tr>
              <tr>
                <td>20201201</td>
                <td>Almaty</td>
                <td>130000</td>
              </tr>
            </table>'''

with open("task4.html", "w") as file:
    file.write(table)

# Convert HTML table to CSV and change date format
with open("task4.html", "r") as in_file, open('task4_fromhtml_out.csv', 'w') as out_file:
    soup = bs(in_file, 'html.parser')
    table = soup.find_all('table')[0]
    for row in table.findAll("tr"):
        cells = row.findAll("td")
    cells_list = [str(tag) for tag in cells]
    pattern = r'<\/?td>'
    cells_clean = [re.sub(pattern, '', item) for item in cells_list]

    writer = csv.writer(out_file)
    ts = datetime.strptime(cells_clean[0], "%Y%m%d")
    cells_clean[0] = ts.strftime('%d-%m-%Y')

    writer.writerow(('date', 'place', 'value'))
    writer.writerow(cells_clean)


# PDF table to CSV

# Create PDF table
pdf_data = (("date", "place", "value"),
            ("20201201", "Almaty", '130000'))
pdf = FPDF()
pdf.add_page()
pdf.set_font("Times", size=16)
with pdf.table() as table:
    for data_row in pdf_data:
        row = table.row()
        for datum in data_row:
            row.cell(datum)
pdf.output('task4.pdf')

# Convert PDF table to CSV and change date format
with open('task4.pdf', 'rb') as pdf_file, open('task4_frompdf_out.csv', 'w') as out_file:
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    writer = csv.writer(out_file)

    table_data = []
    for page_num in range(len(pdf_reader.pages)):
        page = pdf_reader.pages[page_num]
        text = page.extract_text()
        lines = text.split('\n')
        for line in lines:
            row = line.split()
            table_data.append(row)

        ts = datetime.strptime(table_data[1][0], "%Y%m%d")
        table_data[1][0] = ts.strftime('%d-%m-%Y')

        writer.writerows(table_data)
